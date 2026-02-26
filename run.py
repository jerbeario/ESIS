import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd 
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers


def clean_result(result):
    result_map = {'Négatif': '-', 'Postif': '+', '': ''}
    return result_map.get(result, 'N.A.')

def clean_date(date):
    if pd.isna(date):
        return None
    if isinstance(date, pd.Timestamp):
        return date.strftime('%d/%m/%Y')
    parts = date.split('/')
    return f"{parts[0]}/{parts[1]}/{parts[2]}"

def get_age(ddn, year):
    if pd.isna(ddn):
        return None
    age = year - pd.to_datetime(ddn, dayfirst=True).year
    return age - 1

def get_rang(age):
    if age is None:
        return None
    rang = min((age - 50)// 2 + 1, 13) # A verifier 
    return f'R{rang}'

def clean_name(name):
    if pd.isna(name):
        return None
    return name.strip().lower()

def clean_nss(nss):
    if pd.isna(nss):
        return None
    nss = str(nss)
    return int(float(nss.replace(' ', '').replace('|', '')[:13]))

def clean_suivis_df(suivis_df):
    suivis_df['nom'] = suivis_df['nom de naissance'].apply(clean_name)
    suivis_df['prenom'] = suivis_df['prénom'].apply(clean_name)
    suivis_df['ddn'] = suivis_df['date de naissance '].apply(clean_date)
    suivis_df['nss'] = suivis_df['immatriculation sécu entrer les chiffres sans espace (mise en forme spécifique)'].apply(clean_nss)
    return suivis_df

def get_patient_data(input_df, year):
    patient_data = {}
    for index, row in input_df.iterrows():
        nss = int(row['NSS'].replace(' ', '').replace('|', '')[:13])
        nom = clean_name(row['nom_jf'] if not pd.isna(row['nom_jf']) else row['nom'])
        prenom = clean_name(row['prenom'])
        ddn = clean_date(row['ddn'])
        test_date = clean_date(row['date_Realisation_test'])
        test_result = clean_result(row['resultat_test'])
        last_test_date = clean_date(row['date_Test_Avant_Invitation'])
        age = get_age(ddn, year)
        rang = get_rang(age)
        patient_data[nss] = {'age': age, 'test_date': test_date, 'test_result': test_result, 'last_test_date': last_test_date, 'rang': rang, 'nom': nom, 'prenom': prenom, 'ddn': ddn}
    return patient_data

def get_patient_index(suivis_df, patient_nss):
    row = suivis_df.loc[suivis_df['nss'] == patient_nss]
    if row.empty:
        print(f"Patient with NSS {patient_nss} not found in suivis_df.")
        return None
    elif len(row) > 1: 
        print(f"Multiple entries found for NSS {patient_nss}. Please check the data.")
        return None
    print(f"Found patient at index {row.index.values[0] + 2}.")
    return row.index.values[0] + 2 # +2 because of header and 0-indexing

def get_patient_index_from_ddn(suivis_df, name, ddn):
    print(f"Looking for patient with name '{name}' and DDN '{ddn}' in suivis_df...")
    row = suivis_df.loc[(suivis_df['nom'] == name) & (suivis_df['ddn'] == ddn)]
    if row.empty:
        print(f"Patient with name {name} and DDN {ddn} not found in suivis_df.")
        return None
    elif len(row) > 1:
        print(f"Multiple entries found for name {name} and DDN {ddn}. Please check the data.")
        return None
    print(f"Found patient at index {row.index.values[0] + 2}.")
    return row.index.values[0] + 2  # +2 because of header and 0-indexing

def fix_formats(sheet, cols):
    for col in cols:
        fmt = "DD/MM/YYYY"
        for row in sheet.iter_rows(min_row=2, min_col=col, max_col=col, max_row=sheet.max_row):
            cell = row[0]
            if cell.value is not None:
                cell.number_format = fmt

def fill_patient(suivis_excel, suivis_df, patient_data, year):
    sheet = suivis_excel[str(year)]
    print(patient_data)
    for patient_nss, data in patient_data.items():
        row_index = get_patient_index(suivis_df, patient_nss)
        # row_index = get_patient_index_from_ddn(suivis_df, data['nom'], data['ddn'])
        if row_index is None:
            continue

        # change cells
        sheet.cell(row=row_index, column=27).value = data['test_date']
        sheet.cell(row=row_index, column=28).value = data['test_result'] 
        sheet.cell(row=row_index, column=32).value = data['last_test_date']
        sheet.cell(row=row_index, column=31).value = data['rang']

        none_color = PatternFill(start_color="FF215F9A", end_color="FF215F9A", fill_type="solid")
        else_color = PatternFill(start_color="FFA6CAEC", end_color="FFA6CAEC", fill_type="solid")
       
        if data['test_date'] is None:
            sheet.cell(row=row_index, column=27).fill = none_color
        else:
            sheet.cell(row=row_index, column=27).fill = else_color
        
        if data['last_test_date'] is None: 
            sheet.cell(row=row_index, column=32).fill = none_color 
        else: 
            sheet.cell(row=row_index, column=32).fill = else_color
        fix_formats(sheet, [1, 8, 15, 19, 20, 26, 27, 32, 33, 34, 43, 44, 45, 47, 48, 58, 59])


    return suivis_excel


def run_update(input_file, suivis_file, year, overwrite):
    suivis_excel = load_workbook(suivis_file)
    suivis_df = pd.read_excel(suivis_file, sheet_name=str(year))
    input_df = pd.read_csv(input_file, delimiter=';')

    suivis_df = clean_suivis_df(suivis_df)
    patient_data = get_patient_data(input_df, year)
    updated_suvis_excel = fill_patient(suivis_excel, suivis_df, patient_data, year)
    if updated_suvis_excel is None:
        raise ValueError("Update failed due to missing NSS entries. Check console output.")
    if overwrite:
        output_path = suivis_file
    else:
        output_path = suivis_file.parent / f'suivis_patients_{year}_updated.xlsx'
    updated_suvis_excel.save(output_path)
    print(f"Updated file saved to: {output_path}")
    return output_path


def launch_gui():
    root = tk.Tk()
    root.title("ESIS updater")
    root.resizable(False, False)

    input_path_var = tk.StringVar()
    suivis_path_var = tk.StringVar()
    input_label_var = tk.StringVar(value="No file selected")
    suivis_label_var = tk.StringVar(value="No file selected")
    year_var = tk.StringVar()
    overwrite_var = tk.BooleanVar(value=False)

    label_max = 24

    def shorten_label(text, max_len):
        if len(text) <= max_len:
            return text
        keep = max_len - 3
        head = keep // 2
        tail = keep - head
        return f"{text[:head]}...{text[-tail:]}"

    def choose_input():
        path = filedialog.askopenfilename(
            title="Select input file",
            filetypes=[("CSV files", ("*.csv", "*.CSV")), ("All files", "*")],
        )
        if path:
            input_path_var.set(path)
            input_label_var.set(shorten_label(Path(path).name, label_max))

    def choose_suivis():
        path = filedialog.askopenfilename(
            title="Select suivis file",
            filetypes=[
                ("Excel files", ("*.xlsx", "*.xlsm", "*.xls")),
                ("All files", "*"),
            ],
        )
        if path:
            suivis_path_var.set(path)
            suivis_label_var.set(shorten_label(Path(path).name, label_max))

    def run_clicked():
        input_path = input_path_var.get().strip()
        suivis_path = suivis_path_var.get().strip()
        year_text = year_var.get().strip()

        if not input_path or not suivis_path or not year_text:
            messagebox.showerror(
                "Missing data",
                "Please select the input file, the suivis file, and enter a year.",
            )
            return

        try:
            year = int(year_text)
        except ValueError:
            messagebox.showerror("Invalid year", "Year must be a number (e.g., 2025).")
            return

        # try:
        output_path = run_update(
            Path(input_path),
            Path(suivis_path),
            year,
            overwrite_var.get(),
        )
        # except Exception as exc:
        #     messagebox.showerror("Error", f"Failed to update: {exc}")
        #     return

        messagebox.showinfo("Done", f"Saved: {output_path}")
        root.destroy()

    pad = {"padx": 8, "pady": 4}

    tk.Label(root, text="Input file (CSV):").grid(row=0, column=0, sticky="w", **pad)
    tk.Button(root, text="Browse", command=choose_input, width=12).grid(
        row=0, column=1, sticky="w", **pad
    )
    tk.Label(root, textvariable=input_label_var, width=24, anchor="w").grid(
        row=0, column=2, sticky="w", **pad
    )

    tk.Label(root, text="Suivis file (Excel):").grid(row=1, column=0, sticky="w", **pad)
    tk.Button(root, text="Browse", command=choose_suivis, width=12).grid(
        row=1, column=1, sticky="w", **pad
    )
    tk.Label(root, textvariable=suivis_label_var, width=24, anchor="w").grid(
        row=1, column=2, sticky="w", **pad
    )

    tk.Label(root, text="Year:").grid(row=2, column=0, sticky="w", **pad)
    tk.Entry(root, textvariable=year_var, width=12).grid(
        row=2, column=1, sticky="w", **pad
    )

    tk.Checkbutton(
        root,
        text="Overwrite original file",
        variable=overwrite_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=2, column=2, columnspan=2, sticky="w", **pad)

    tk.Button(root, text="Quit", command=root.destroy, width=12).grid(
        row=4, column=1, sticky="w", **pad
    )
    tk.Button(root, text="Run", command=run_clicked, width=12).grid(
        row=4, column=1, sticky="e", **pad
    )

    root.mainloop()


def main():
    launch_gui()

if __name__ == "__main__":
    main()
